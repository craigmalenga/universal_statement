# backend/app/export.py
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import logging
from typing import Tuple

logger = logging.getLogger(__name__)

def export_to_files(df: pd.DataFrame, session_id: str, temp_dir: Path) -> Tuple[str, str]:
    """
    Export DataFrame to both Excel and CSV formats
    """
    try:
        # Prepare data for export
        export_df = prepare_data_for_export(df)
        
        # Generate file paths
        excel_path = temp_dir / f"{session_id}_transactions.xlsx"
        csv_path = temp_dir / f"{session_id}_transactions.csv"
        
        # Export to Excel with formatting
        export_to_excel(export_df, str(excel_path))
        
        # Export to CSV
        export_to_csv(export_df, str(csv_path))
        
        logger.info(f"Successfully exported {len(export_df)} transactions to Excel and CSV")
        
        return str(excel_path), str(csv_path)
        
    except Exception as e:
        logger.error(f"Error exporting files: {str(e)}")
        raise

def prepare_data_for_export(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare DataFrame for export with proper formatting
    """
    if df.empty:
        return pd.DataFrame(columns=['Date', 'Description', 'Debit', 'Credit', 'Balance'])
    
    # Create a copy for export
    export_df = df.copy()
    
    # Rename columns to proper case
    column_mapping = {
        'date': 'Date',
        'description': 'Description',
        'debit': 'Debit',
        'credit': 'Credit',
        'balance': 'Balance'
    }
    export_df = export_df.rename(columns=column_mapping)
    
    # Ensure all required columns exist
    required_columns = ['Date', 'Description', 'Debit', 'Credit', 'Balance']
    for col in required_columns:
        if col not in export_df.columns:
            export_df[col] = None
    
    # Reorder columns
    export_df = export_df[required_columns]
    
    # Format amounts to 2 decimal places
    for col in ['Debit', 'Credit', 'Balance']:
        export_df[col] = export_df[col].apply(lambda x: f"{x:.2f}" if pd.notnull(x) else "")
    
    # Clean up description
    export_df['Description'] = export_df['Description'].fillna('').astype(str)
    export_df['Description'] = export_df['Description'].apply(lambda x: x.strip())
    
    # Replace NaN values with empty strings for better display
    export_df = export_df.fillna('')
    
    return export_df

def export_to_excel(df: pd.DataFrame, file_path: str):
    """
    Export DataFrame to Excel with professional formatting
    """
    try:
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bank Transactions"
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Apply formatting
        format_excel_worksheet(ws, len(df))
        
        # Save the workbook
        wb.save(file_path)
        
        # Verify the file was created and is valid
        verify_excel_file(file_path)
        
        logger.info(f"Excel file exported successfully: {file_path}")
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        raise

def format_excel_worksheet(ws, data_rows: int):
    """
    Apply professional formatting to Excel worksheet
    """
    try:
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        data_font = Font(name='Arial', size=11)
        data_alignment = Alignment(horizontal='left', vertical='center')
        amount_alignment = Alignment(horizontal='right', vertical='center')
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Format data rows
        for row in range(2, data_rows + 2):
            for col in range(1, 6):  # 5 columns
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                
                # Right-align amount columns
                if col in [3, 4, 5]:  # Debit, Credit, Balance columns
                    cell.alignment = amount_alignment
                    # Format as currency if it contains a number
                    if cell.value and str(cell.value).replace('.', '').replace('-', '').isdigit():
                        cell.number_format = '£#,##0.00'
                else:
                    cell.alignment = data_alignment
        
        # Auto-adjust column widths
        column_widths = {
            'A': 12,  # Date
            'B': 40,  # Description
            'C': 15,  # Debit
            'D': 15,  # Credit
            'E': 15,  # Balance
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        # Add alternating row colors for better readability
        light_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        for row in range(3, data_rows + 2, 2):  # Every other row starting from row 3
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = light_fill
        
    except Exception as e:
        logger.warning(f"Error formatting Excel worksheet: {str(e)}")

def verify_excel_file(file_path: str):
    """
    Verify that the Excel file was created correctly and can be opened
    """
    try:
        # Try to open and read the file
        test_wb = openpyxl.load_workbook(file_path)
        test_ws = test_wb.active
        
        # Check that it has data
        if test_ws.max_row < 1:
            raise Exception("Excel file appears to be empty")
        
        # Check that headers exist
        headers = [cell.value for cell in test_ws[1]]
        required_headers = ['Date', 'Description', 'Debit', 'Credit', 'Balance']
        
        for header in required_headers:
            if header not in headers:
                raise Exception(f"Missing required header: {header}")
        
        test_wb.close()
        logger.info("Excel file verification successful")
        
    except Exception as e:
        logger.error(f"Excel file verification failed: {str(e)}")
        raise Exception(f"Created Excel file is invalid: {str(e)}")

def export_to_csv(df: pd.DataFrame, file_path: str):
    """
    Export DataFrame to CSV format
    """
    try:
        # Export with proper encoding and formatting
        df.to_csv(
            file_path,
            index=False,
            encoding='utf-8-sig',  # BOM for Excel compatibility
            float_format='%.2f',
            date_format='%Y-%m-%d'
        )
        
        # Verify the file was created
        verify_csv_file(file_path, len(df))
        
        logger.info(f"CSV file exported successfully: {file_path}")
        
    except Exception as e:
        logger.error(f"Error exporting to CSV: {str(e)}")
        raise

def verify_csv_file(file_path: str, expected_rows: int):
    """
    Verify that the CSV file was created correctly
    """
    try:
        # Read the file back to verify
        test_df = pd.read_csv(file_path, encoding='utf-8-sig')
        
        # Check row count (excluding header)
        if len(test_df) != expected_rows:
            logger.warning(f"CSV row count mismatch: expected {expected_rows}, got {len(test_df)}")
        
        # Check required columns
        required_columns = ['Date', 'Description', 'Debit', 'Credit', 'Balance']
        missing_columns = [col for col in required_columns if col not in test_df.columns]
        
        if missing_columns:
            raise Exception(f"Missing columns in CSV: {missing_columns}")
        
        logger.info("CSV file verification successful")
        
    except Exception as e:
        logger.error(f"CSV file verification failed: {str(e)}")
        raise Exception(f"Created CSV file is invalid: {str(e)}")

def create_summary_sheet(wb, df: pd.DataFrame):
    """
    Create a summary sheet with transaction statistics
    """
    try:
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary")
        
        # Calculate statistics
        total_debits = df['Debit'].dropna().sum() if 'Debit' in df.columns else 0
        total_credits = df['Credit'].dropna().sum() if 'Credit' in df.columns else 0
        transaction_count = len(df)
        date_range = f"{df['Date'].min()} to {df['Date'].max()}" if 'Date' in df.columns else "N/A"
        
        # Add summary data
        summary_data = [
            ["Bank Statement Summary", ""],
            ["", ""],
            ["Total Transactions", transaction_count],
            ["Date Range", date_range],
            ["Total Debits", f"£{total_debits:.2f}"],
            ["Total Credits", f"£{total_credits:.2f}"],
            ["Net Amount", f"£{total_credits - total_debits:.2f}"],
        ]
        
        for row_data in summary_data:
            summary_ws.append(row_data)
        
        # Format summary sheet
        summary_ws['A1'].font = Font(name='Arial', size=16, bold=True)
        summary_ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        summary_ws['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        
        # Auto-adjust column widths
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 20
        
    except Exception as e:
        logger.warning(f"Error creating summary sheet: {str(e)}")